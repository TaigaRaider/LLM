import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import or_
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import Officer, Participant
from ..schemas import ImportResult, ParticipantIn, ParticipantOut, ParticipantUpdate
from ..permissions import require_perms

router = APIRouter(prefix="/api/participants", tags=["participants"])


def _next_id_number(db: Session) -> str:
    max_num = 0
    for p in db.query(Participant).all():
        digits = "".join(ch for ch in p.id_number if ch.isdigit())
        if digits.isdigit():
            max_num = max(max_num, int(digits))
    return f"ID-{max_num + 1:04d}"


@router.get("", response_model=list[ParticipantOut])
def list_participants(
    search: str = "",
    active_only: bool = False,
    include_inactive: bool = False,
    db: Session = Depends(get_db),
    _: Officer = Depends(require_perms("admin")),
):
    q = db.query(Participant)
    if search:
        s = f"%{search.strip().lower()}%"
        q = q.filter(
            or_(
                Participant.name.ilike(s),
                Participant.id_number.ilike(s),
                Participant.phone.ilike(s),
            )
        )
    if active_only:
        q = q.filter(Participant.active.is_(True))
    elif not include_inactive:
        q = q.filter(Participant.active.is_(True))
    return q.order_by(Participant.name).all()


@router.post("", response_model=ParticipantOut, status_code=201)
def create_participant(
    body: ParticipantIn,
    db: Session = Depends(get_db),
    _: Officer = Depends(require_perms("admin")),
):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=422, detail="Name is required")
    id_number = body.id_number.strip() or _next_id_number(db)
    existing = db.query(Participant).filter(Participant.id_number == id_number).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"ID number {id_number} already registered")
    participant = Participant(
        name=name,
        id_number=id_number,
        phone=body.phone.strip(),
        group=body.group.strip(),
        external_id=body.external_id,
        source="local",
    )
    db.add(participant)
    db.commit()
    db.refresh(participant)
    return participant


@router.post("/bulk", response_model=list[ParticipantOut], status_code=201)
def bulk_create_participants(
    items: list[ParticipantIn],
    db: Session = Depends(get_db),
    _: Officer = Depends(require_perms("admin")),
):
    created = []
    existing_numbers = {p.id_number for p in db.query(Participant).all()}
    for item in items:
        name = item.name.strip()
        if not name:
            continue
        id_number = item.id_number.strip() or _next_id_number(db)
        if id_number in existing_numbers:
            continue
        participant = Participant(
            name=name,
            id_number=id_number,
            phone=item.phone.strip(),
            group=item.group.strip(),
            external_id=item.external_id,
            source="local",
        )
        db.add(participant)
        existing_numbers.add(id_number)
        created.append(participant)
    db.commit()
    for p in created:
        db.refresh(p)
    return created


@router.post("/import-excel", response_model=ImportResult)
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: Officer = Depends(require_perms("admin")),
):
    # Validate file extension
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=422, detail="File must be .xlsx or .xlsm")

    # Read file with size limit (5 MB)
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=422, detail="File too large (max 5 MB)")

    # Parse Excel
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not parse Excel file: {exc}")

    if not raw_rows:
        raise HTTPException(status_code=422, detail="The file has no rows")

    # Detect header row
    header = [str(c or "").strip().lower().replace(" ", "").replace("_", "") for c in raw_rows[0]]
    idx_name = next((i for i, h in enumerate(header) if h in ("name", "fullname", "participantname")), None)
    idx_id = next((i for i, h in enumerate(header) if h in ("idnumber", "id", "nationalid", "participantid")), None)
    idx_phone = next((i for i, h in enumerate(header) if h in ("phone", "phonenumber", "mobile")), None)
    idx_group = next((i for i, h in enumerate(header) if h in ("group", "bus", "busgroup")), None)

    if idx_name is None:
        raise HTTPException(status_code=422, detail="Missing required column: 'name' (or 'fullname', 'participantname')")

    existing_numbers = {p.id_number for p in db.query(Participant).all()}
    created = 0
    skipped = []
    total = max(len(raw_rows) - 1, 0)

    for row_idx, row in enumerate(raw_rows[1:], start=2):
        if not row or not any(str(c or "").strip() for c in row):
            continue
        cell = lambda i: str(row[i]).strip() if i is not None and i < len(row) and row[i] is not None else ""
        name = cell(idx_name)
        if not name:
            skipped.append(f"Row {row_idx}: missing name")
            continue
        id_number = cell(idx_id) if idx_id is not None else ""
        if not id_number:
            id_number = _next_id_number(db)
        if id_number in existing_numbers:
            skipped.append(f"Row {row_idx}: duplicate ID '{id_number}'")
            continue
        existing_numbers.add(id_number)
        participant = Participant(
            name=name,
            id_number=id_number,
            phone=cell(idx_phone),
            group=cell(idx_group),
            source="local",
        )
        db.add(participant)
        created += 1

    db.commit()
    return ImportResult(total=total, created=created, skipped=skipped[:100])


@router.put("/{participant_id}", response_model=ParticipantOut)
def update_participant(
    participant_id: str,
    body: ParticipantUpdate,
    db: Session = Depends(get_db),
    _: Officer = Depends(require_perms("admin")),
):
    participant = db.get(Participant, participant_id)
    if not participant:
        raise HTTPException(status_code=404, detail="Participant not found")
    if body.name is not None:
        participant.name = body.name.strip()
    if body.id_number is not None:
        new_number = body.id_number.strip()
        if new_number:
            clash = db.query(Participant).filter(
                Participant.id_number == new_number, Participant.id != participant_id
            ).first()
            if clash:
                raise HTTPException(status_code=409, detail=f"ID number {new_number} already registered")
            participant.id_number = new_number
    if body.phone is not None:
        participant.phone = body.phone.strip()
    if body.group is not None:
        participant.group = body.group.strip()
    if body.active is not None:
        participant.active = body.active
    db.commit()
    db.refresh(participant)
    return participant


@router.delete("/{participant_id}", response_model=ParticipantOut)
def delete_participant(
    participant_id: str,
    db: Session = Depends(get_db),
    _: Officer = Depends(require_perms("admin")),
):
    participant = db.get(Participant, participant_id)
    if not participant:
        raise HTTPException(status_code=404, detail="Participant not found")
    participant.active = False
    db.commit()
    db.refresh(participant)
    return participant